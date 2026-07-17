"""
app.py — Flask web server + REST API for the Receipt Scanner.
Run: python app.py
"""

import csv
import io
import logging
import os
import threading
from datetime import datetime
from pathlib import Path

from flask import (
    Flask,
    abort,
    jsonify,
    request,
    send_file,
    send_from_directory,
    render_template,
)

import base64
import secrets

import database
import drive_watcher
import processor
import watcher
from config import GOOGLE_DRIVE_FOLDER_ID, ORIGINALS_DIR, THUMBNAILS_DIR, WATCH_FOLDER, CATEGORIES

# ── Access control (opt-in via env) ──────────────────────────────────────────
# APP_PASSWORD:    when set, every request needs HTTP Basic auth (any username).
#                  REQUIRED for a public deploy — without it, all receipts,
#                  card digits, and spend data are open to the internet.
# AGENT_API_TOKEN: when set, requests may instead present
#                  "Authorization: Bearer <token>" (used by Claude automations).
# When neither is set (local dev), no auth is enforced.
APP_PASSWORD = os.environ.get("APP_PASSWORD", "")
AGENT_API_TOKEN = os.environ.get("AGENT_API_TOKEN", "")

def _parse_multi(val, sep="|"):
    """Parse a pipe-separated query param into a list, or None."""
    if not val:
        return None
    items = [v.strip() for v in val.split(sep) if v.strip()]
    return items or None


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

@app.before_request
def _require_auth():
    if not APP_PASSWORD and not AGENT_API_TOKEN:
        return None  # local dev, auth disabled
    header = request.headers.get("Authorization", "")
    if AGENT_API_TOKEN and header == f"Bearer {AGENT_API_TOKEN}":
        return None
    if APP_PASSWORD and header.startswith("Basic "):
        try:
            decoded = base64.b64decode(header.split(" ", 1)[1]).decode("utf-8")
            _user, _, pw = decoded.partition(":")
            if secrets.compare_digest(pw, APP_PASSWORD):
                return None
        except Exception:
            pass
    resp = jsonify({"error": "authentication required"})
    resp.status_code = 401
    resp.headers["WWW-Authenticate"] = 'Basic realm="Receipt Scanner"'
    return resp

app.config["JSON_SORT_KEYS"] = False


# ── Startup ──────────────────────────────────────────────────────────────────

@app.before_request
def _startup():
    # Flask calls this once per request, guard with a flag
    pass


def _init():
    database.init_db()
    if GOOGLE_DRIVE_FOLDER_ID:
        drive_watcher.start()
        logger.info("Ingestion mode: Google Drive API (folder %s)", GOOGLE_DRIVE_FOLDER_ID)
    else:
        watcher.start()
        logger.info("Ingestion mode: local folder watchdog (%s)", WATCH_FOLDER)
    logger.info("Receipt Scanner ready at http://127.0.0.1:5000")


# ── Pages ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/receipt/<int:rid>")
def receipt_detail_page(rid):
    return render_template("detail.html", receipt_id=rid)


# ── Static assets (receipts) ─────────────────────────────────────────────────

@app.route("/receipts/originals/<path:filename>")
def serve_original(filename):
    return send_from_directory(ORIGINALS_DIR, filename)


@app.route("/receipts/thumbnails/<path:filename>")
def serve_thumbnail(filename):
    return send_from_directory(THUMBNAILS_DIR, filename)


# ── API: Stats ───────────────────────────────────────────────────────────────

@app.route("/api/agent/summary")
def api_agent_summary():
    """Compact read-only summary for Claude automations (Monday brief,
    Command Center, QuickBooks reconciliation)."""
    stats = database.get_stats()
    return jsonify({
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "stats": stats,
        "monthly_category_totals": database.get_monthly_category_totals(months=3),
        "recent_receipts": database.get_recent_receipts(limit=25),
    })


@app.route("/api/stats")
def api_stats():
    fid = request.args.get("folder_id")
    filter_keys = ["search", "cards", "vendors", "category", "date_from", "date_to",
                   "folder_id", "missing", "show_hidden"]
    has_filters = any(request.args.get(k) for k in filter_keys)
    if has_filters:
        stats = database.get_filtered_stats(
            search=request.args.get("search"),
            cards=_parse_multi(request.args.get("cards")),
            vendors=_parse_multi(request.args.get("vendors")),
            category=request.args.get("category"),
            date_from=request.args.get("date_from"),
            date_to=request.args.get("date_to"),
            folder_id=int(fid) if fid else None,
            show_hidden=request.args.get("show_hidden") == "1",
            missing=request.args.get("missing"),
        )
        global_stats = database.get_stats()
        stats["pending"] = global_stats["pending"]
        stats["review"]  = global_stats["review"]
        return jsonify(stats)
    return jsonify(database.get_stats())


# ── API: Receipts ────────────────────────────────────────────────────────────

@app.route("/api/receipts")
def api_list_receipts():
    fid = request.args.get("folder_id")
    rows, total = database.list_receipts(
        search=request.args.get("search"),
        cards=_parse_multi(request.args.get("cards")),
        vendors=_parse_multi(request.args.get("vendors")),
        category=request.args.get("category"),
        date_from=request.args.get("date_from"),
        date_to=request.args.get("date_to"),
        sort_by=request.args.get("sort_by", "date"),
        sort_dir=request.args.get("sort_dir", "desc"),
        page=int(request.args.get("page", 1)),
        per_page=int(request.args.get("per_page", 50)),
        show_hidden=request.args.get("show_hidden") == "1",
        folder_id=int(fid) if fid else None,
        missing=request.args.get("missing"),
    )
    return jsonify({"receipts": rows, "total": total})


@app.route("/api/receipts/<int:rid>")
def api_get_receipt(rid):
    receipt = database.get_receipt(rid)
    if not receipt:
        abort(404)
    items = database.get_items(rid)
    receipt["items"] = items
    return jsonify(receipt)


@app.route("/api/receipts/<int:rid>", methods=["PUT"])
def api_update_receipt(rid):
    receipt = database.get_receipt(rid)
    if not receipt:
        abort(404)

    body = request.get_json(force=True) or {}
    allowed = {
        "vendor_name", "date", "card_last4", "card_type",
        "total_amount", "category", "notes",
    }
    update = {k: v for k, v in body.items() if k in allowed}
    if update:
        database.update_receipt(rid, update)

    if "items" in body:
        database.replace_items(rid, body["items"])

    receipt = database.get_receipt(rid)
    receipt["items"] = database.get_items(rid)
    return jsonify(receipt)


@app.route("/api/receipts/<int:rid>", methods=["DELETE"])
def api_delete_receipt(rid):
    receipt = database.get_receipt(rid)
    if not receipt:
        abort(404)

    # Remove image files
    for path_key in ("stored_path", "thumbnail_path"):
        p = receipt.get(path_key)
        if p and os.path.isfile(p):
            try:
                os.remove(p)
            except OSError:
                pass

    database.delete_receipt(rid)
    return jsonify({"ok": True})


@app.route("/api/receipts/<int:rid>/hide", methods=["POST"])
def api_hide_receipt(rid):
    receipt = database.get_receipt(rid)
    if not receipt:
        abort(404)
    database.update_receipt(rid, {"status": "hidden"})
    return jsonify({"ok": True})


@app.route("/api/receipts/<int:rid>/unhide", methods=["POST"])
def api_unhide_receipt(rid):
    receipt = database.get_receipt(rid)
    if not receipt:
        abort(404)
    database.update_receipt(rid, {"status": "processed"})
    return jsonify({"ok": True})


# ── API: Cards & Categories ───────────────────────────────────────────────────

@app.route("/api/cards")
def api_cards():
    return jsonify(database.get_cards())


@app.route("/api/cards/merge", methods=["POST"])
def api_merge_card():
    body = request.get_json(silent=True) or {}
    card_last4 = body.get("card_last4", "").strip()
    card_type  = body.get("card_type", "").strip() or None
    if not card_last4:
        return jsonify({"error": "card_last4 required"}), 400
    updated = database.merge_card(card_last4, card_type)
    return jsonify({"ok": True, "updated": updated})


@app.route("/api/vendors")
def api_vendors():
    return jsonify(database.get_vendors())


@app.route("/api/categories")
def api_categories():
    return jsonify(CATEGORIES)


# ── API: Folders ─────────────────────────────────────────────────────────────

@app.route("/api/folders")
def api_list_folders():
    return jsonify(database.get_folders())


@app.route("/api/folders", methods=["POST"])
def api_create_folder():
    body = request.get_json(silent=True) or {}
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name required"}), 400
    fid = database.create_folder(
        name=name,
        parent_id=body.get("parent_id"),
        color=body.get("color", "#6366f1"),
    )
    return jsonify({"id": fid, "name": name})


@app.route("/api/folders/<int:fid>", methods=["PUT"])
def api_update_folder(fid):
    body = request.get_json(silent=True) or {}
    database.update_folder(fid, body)
    return jsonify({"ok": True})


@app.route("/api/folders/<int:fid>", methods=["DELETE"])
def api_delete_folder(fid):
    database.delete_folder(fid)
    return jsonify({"ok": True})


@app.route("/api/receipts/<int:rid>/folders")
def api_get_receipt_folders(rid):
    return jsonify(database.get_receipt_folder_ids(rid))


@app.route("/api/receipts/<int:rid>/folders", methods=["PUT"])
def api_set_receipt_folders(rid):
    body = request.get_json(silent=True) or {}
    folder_ids = body.get("folder_ids", [])
    database.set_receipt_folders(rid, folder_ids)
    return jsonify({"ok": True})


@app.route("/api/folders/<int:fid>/tag-results", methods=["POST"])
def api_bulk_tag(fid):
    """Tag all receipts matching the posted filter params with a folder."""
    body = request.get_json(silent=True) or {}
    receipt_ids = database.get_all_receipt_ids_for_filter(
        search=body.get("search"),
        cards=body.get("cards"),   # already a list from JS
        vendors=body.get("vendors"),
        category=body.get("category"),
        date_from=body.get("date_from"),
        date_to=body.get("date_to"),
        folder_id=int(body["folder_id"]) if body.get("folder_id") else None,
    )
    database.bulk_tag_folder(fid, receipt_ids)
    return jsonify({"ok": True, "tagged": len(receipt_ids)})


# ── API: Bulk operations ──────────────────────────────────────────────────────

@app.route("/api/receipts/bulk-update", methods=["POST"])
def api_bulk_update_receipts():
    body = request.get_json(silent=True) or {}
    ids = body.get("ids", [])
    data = body.get("data", {})
    if not ids or not data:
        return jsonify({"error": "ids and data required"}), 400
    database.bulk_update_receipts(ids, data)
    return jsonify({"ok": True, "updated": len(ids)})


@app.route("/api/receipts/bulk-tag", methods=["POST"])
def api_bulk_tag_receipts():
    body = request.get_json(silent=True) or {}
    ids = body.get("ids", [])
    folder_id = body.get("folder_id")
    if not ids or not folder_id:
        return jsonify({"error": "ids and folder_id required"}), 400
    database.bulk_tag_folder(int(folder_id), ids)
    return jsonify({"ok": True, "tagged": len(ids)})


# ── API: Review Queue ─────────────────────────────────────────────────────────

@app.route("/api/review")
def api_review_queue():
    rows = database.list_review_queue()
    return jsonify(rows)


@app.route("/api/review/<int:rid>/approve", methods=["POST"])
def api_review_approve(rid):
    """Move a review item into the main receipt list."""
    receipt = database.get_receipt(rid)
    if not receipt:
        abort(404)
    if receipt["status"] != "review":
        return jsonify({"error": "Receipt is not in review queue"}), 400
    database.update_receipt(rid, {"status": "processed"})
    return jsonify({"ok": True})


@app.route("/api/review/<int:rid>/dismiss", methods=["POST"])
def api_review_dismiss(rid):
    """Dismiss a review item — removes it and its image files."""
    receipt = database.get_receipt(rid)
    if not receipt:
        abort(404)
    if receipt["status"] != "review":
        return jsonify({"error": "Receipt is not in review queue"}), 400
    for path_key in ("stored_path", "thumbnail_path"):
        p = receipt.get(path_key)
        if p and os.path.isfile(p):
            try:
                os.remove(p)
            except OSError:
                pass
    database.delete_receipt(rid)
    return jsonify({"ok": True})


# ── API: Watcher / Scan ───────────────────────────────────────────────────────

@app.route("/api/status")
def api_status():
    status = processor.get_status()
    if GOOGLE_DRIVE_FOLDER_ID:
        status["watching"] = drive_watcher.is_running()
        status["watch_folder"] = f"Google Drive folder {GOOGLE_DRIVE_FOLDER_ID}"
        status["drive"] = drive_watcher.get_status()
    else:
        status["watching"] = watcher.is_running()
        status["watch_folder"] = watcher.get_watch_folder()
    return jsonify(status)


@app.route("/api/scan", methods=["POST"])
def api_scan():
    """Trigger a manual scan (Drive poll in drive mode, folder scan otherwise)."""
    body = request.get_json(silent=True) or {}

    if GOOGLE_DRIVE_FOLDER_ID and not body.get("folder"):
        def _run_drive():
            results = drive_watcher.poll_once()
            logger.info("Manual Drive poll complete: %s", results)

        threading.Thread(target=_run_drive, daemon=True).start()
        return jsonify({"ok": True, "folder": f"Google Drive folder {GOOGLE_DRIVE_FOLDER_ID}"})

    target = body.get("folder") or watcher.get_watch_folder()

    def _run():
        results = processor.scan_folder(target)
        logger.info("Manual scan complete: %s", results)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "folder": target})


@app.route("/api/watch-folder", methods=["POST"])
def api_set_watch_folder():
    body = request.get_json(force=True) or {}
    folder = body.get("folder", "").strip()
    if not folder:
        return jsonify({"error": "folder required"}), 400
    if not os.path.isdir(folder):
        return jsonify({"error": f"Folder not found: {folder}"}), 400
    watcher.set_watch_folder(folder)
    return jsonify({"ok": True, "folder": folder})


# ── API: Export ───────────────────────────────────────────────────────────────

def _build_export_rows(rows):
    """Flatten receipts to CSV/XLSX rows."""
    out = []
    for r in rows:
        out.append(
            {
                "ID": r["id"],
                "Vendor": r.get("vendor_name") or "",
                "Date": r.get("date") or "",
                "Amount": r.get("total_amount") or "",
                "Card Type": r.get("card_type") or "",
                "Card Last 4": r.get("card_last4") or "",
                "Category": r.get("category") or "",
                "Items Preview": r.get("items_preview") or "",
            }
        )
    return out


def _get_filtered_rows():
    fid = request.args.get("folder_id")
    rows, _ = database.list_receipts(
        search=request.args.get("search"),
        cards=_parse_multi(request.args.get("cards")),
        vendors=_parse_multi(request.args.get("vendors")),
        category=request.args.get("category"),
        date_from=request.args.get("date_from"),
        date_to=request.args.get("date_to"),
        sort_by=request.args.get("sort_by", "date"),
        sort_dir=request.args.get("sort_dir", "desc"),
        per_page=10000,
        folder_id=int(fid) if fid else None,
        missing=request.args.get("missing"),
    )
    return rows


@app.route("/api/export/csv")
def api_export_csv():
    rows = _get_filtered_rows()
    export_rows = _build_export_rows(rows)
    buf = io.StringIO()
    if export_rows:
        writer = csv.DictWriter(buf, fieldnames=export_rows[0].keys())
        writer.writeheader()
        writer.writerows(export_rows)
    buf.seek(0)
    return send_file(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"receipts_{datetime.now().strftime('%Y%m%d')}.csv",
    )


@app.route("/api/export/xlsx")
def api_export_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = _get_filtered_rows()
    export_rows = _build_export_rows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Receipts"

    if export_rows:
        headers = list(export_rows[0].keys())
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="6366F1")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in export_rows:
            ws.append(list(row.values()))

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"receipts_{datetime.now().strftime('%Y%m%d')}.xlsx",
    )


@app.route("/api/export/pdf")
def api_export_pdf():
    from fpdf import FPDF

    rows = _get_filtered_rows()

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Receipt Export", ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True, align="C")
    pdf.ln(4)

    headers = ["ID", "Vendor", "Date", "Amount", "Card", "Category"]
    col_w = [15, 70, 25, 25, 30, 30]

    pdf.set_fill_color(99, 102, 241)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    for h, w in zip(headers, col_w):
        pdf.cell(w, 8, h, border=1, fill=True)
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)
    fill = False
    for r in rows:
        pdf.set_fill_color(243, 244, 246) if fill else pdf.set_fill_color(255, 255, 255)
        card = ""
        if r.get("card_type") and r.get("card_last4"):
            card = f"{r['card_type']} -{r['card_last4']}"
        amt = f"${r['total_amount']:.2f}" if r.get("total_amount") is not None else ""
        vals = [
            str(r["id"]),
            (r.get("vendor_name") or "")[:40],
            r.get("date") or "",
            amt,
            card,
            r.get("category") or "",
        ]
        for v, w in zip(vals, col_w):
            pdf.cell(w, 7, v, border=1, fill=True)
        pdf.ln()
        fill = not fill

    buf = io.BytesIO(pdf.output())
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"receipts_{datetime.now().strftime('%Y%m%d')}.pdf",
    )


# ── Run ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    _init()
    app.run(host="127.0.0.1", port=5000, debug=False, use_reloader=False)
